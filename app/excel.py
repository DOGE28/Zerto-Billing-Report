from openpyxl import Workbook, load_workbook
from datetime import datetime
import zerto as z

# Function to flatten and structure the data
def flatten_data(data):
    flat_data = []
    zorg_name = data['zorgName']
    vpg_name = data['VPG']['VpgName']
    
    # Loop through each VM and collect relevant information
    for vm in data['VPG']['VMs']:
        row = [
            datetime.now().strftime("%Y-%m-%d"),  # Timestamp
            zorg_name,  # Zorg name
            vpg_name,  # VPG name
            vm['VmName'],  # VM name
            vm['Cpu']['NumberOfvCpus'],  # Number of CPUs
            vm['Cpu']['CpuUsedInMhz'],  # CPU used in MHz
            vm['Memory']['ActiveGuestMemoryInMB'],  # Active Guest Memory
            vm['Memory']['ConsumedHostMemoryInMB'],  # Consumed Host Memory
            vm['Memory']['MemoryInMB'],  # Total Memory
            vm['Storage']['VolumesProvisionedStorageInGB'],  # Provisioned Storage in GB
            vm['Storage']['VolumesUsedStorageInGB'],  # Used Storage in GB
        ]
        flat_data.append(row)
    
    return flat_data

# Data sample (replace this with the actual data from your script)
# data = {
#     'zorgName': 'Journal Tech',
#     'VPG': {
#         'VpgName': 'riverside-test',
#         'VMs': [
#             {
#                 'VmName': 'api-test01',
#                 'Cpu': {
#                     'NumberOfvCpus': 4,
#                     'CpuUsedInMhz': 269
#                 },
#                 'Memory': {
#                     'ActiveGuestMemoryInMB': 737,
#                     'ConsumedHostMemoryInMB': 8297,
#                     'MemoryInMB': 8192,
#                 },
#                 'Storage': {
#                     'NumberOfVolumes': 1,
#                     'VolumesProvisionedStorageInGB': 88.15,
#                     'VolumesUsedStorageInGB': 71.499
#                 }
#             }
#         ]
#     }
# }

zerto = z.ZertoGet()
zorgs = zerto.get_zorgs_by_vpg()

data = []
for zorg in zorgs:
    resources = zerto.get_zorg_info_from_resources(zorg)
    for resource in resources:
        data.append(resource)

for dat in data:



    # Flatten the data
    flat_data = flatten_data(dat)

    # Excel file path
    excel_file = "new_vms_data.xlsx"

    # Try to load an existing workbook, else create a new one
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        
        # Write the header row
        headers = [
            "Timestamp", "Zorg Name", "VPG Name", "VM Name", "vCPUs", 
            "CPU Used (MHz)", "Active Guest Memory (MB)", "Consumed Host Memory (MB)",
            "Total Memory (MB)", "Provisioned Storage (GB)", "Used Storage (GB)"
        ]
        sheet.append(headers)

    # Append new data
    for row in flat_data:
        sheet.append(row)

    # Save the workbook
    workbook.save(excel_file)

    print("Data added to Excel!")




























# import pandas as pd
# import zerto as z

# zerto = z.ZertoGet()

# data = {}

# def data_zorgs():
#     zorg_done = []
#     zorgs = zerto.get_zorgs_by_vpg()
#     for zorg in zorgs:
#         resources = zerto.get_zorg_info_from_resources(zorg)
#         for resource in resources:
#             if resource['zorg_name'] not in zorg_done:
#                 zorg_done.append(resource['zorg_name'])
#             else:
#                 continue
#     data['Zorg'] = zorg_done


# def data_vpgs_count():
#     vpgs_done = []
#     zorgs = zerto.get_zorgs_by_vpg()
#     for zorg in zorgs:
#         resources = zerto.get_zorg_info_from_resources(zorg)
#         count = 0
#         for resource in resources:
#             count += 1
#         if count == 0:
#             continue
#         else:
#             vpgs_done.append(count)
#     data["VPGs"]= vpgs_done

# def data_vms_count():
#     vms_done = []
#     zorgs = zerto.get_zorgs_by_vpg()
#     for zorg in zorgs:
#         resources = zerto.get_zorg_info_from_resources(zorg)
#         count = 0
#         for resource in resources:
#             for vm in resource['VPG']['VMs']:
#                 count += 1
#         if count == 0:
#             continue
#         else:
#             vms_done.append(count)
#     data["VMs"] = vms_done

# data_zorgs()
# data_vpgs_count()
# data_vms_count()

# print(data)
                  

# df = pd.DataFrame(data)
# df.to_excel('zorgs.xlsx', index=False)
